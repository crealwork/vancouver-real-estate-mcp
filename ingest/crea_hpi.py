"""Parse CREA's national MLS(R) HPI archive — the current-era board layer.

Source: https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/
which publishes a monthly zip of four workbooks (monthly seasonally adjusted
and not, plus quarterly and annual not-adjusted).

Covers 2005-01 to the current month at board level. We keep the boards that
make up Metro Vancouver + Fraser Valley, plus BC and national aggregates so
an agent can answer "versus the province / the country" without a second
source.
"""

from __future__ import annotations

import re
import sys
import zipfile

import openpyxl

from common import OUT, RAW, UA, DecodeError, num, slugify, write_jsonl

HPI_TOOL_URL = "https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/"

# Board-level sheets relevant to a Metro Vancouver + Fraser Valley server.
SHEETS = {
    "GREATER_VANCOUVER": "Greater Vancouver (board)",
    "FRASER_VALLEY": "Fraser Valley (board)",
    "LOWER_MAINLAND": "Lower Mainland (GV+FV)",
    "CHILLIWACK_AND_DISTRICT": "Chilliwack & District (board)",
    "BRITISH_COLUMBIA": "British Columbia",
    "AGGREGATE": "Canada",
}

# CREA's column names -> our property-type vocabulary.
TYPES = {
    "Composite": "composite",
    "Single_Family": "detached",
    "One_Storey": "detached_one_storey",
    "Two_Storey": "detached_two_storey",
    "Townhouse": "townhouse",
    "Apartment": "apartment",
}

WORKBOOKS = {
    "Not Seasonally Adjusted (M).xlsx": ("monthly", False),
    "Seasonally Adjusted (M).xlsx": ("monthly", True),
    "Not Seasonally Adjusted (Q).xlsx": ("quarterly", False),
    "Not Seasonally Adjusted (A).xlsx": ("annual", False),
}


def find_latest_zip_url() -> str:
    """Scrape the HPI tool page for the current month's archive URL."""
    import urllib.request

    req = urllib.request.Request(HPI_TOOL_URL, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as resp:
        html = resp.read().decode("utf-8", "replace")
    urls = re.findall(r'https://www\.crea\.ca/files/mls-hpi-data/MLS_HPI-[^"\']+_EN\.zip', html)
    if not urls:
        raise DecodeError("no MLS_HPI zip link found on the CREA HPI tool page")
    return urls[0]


def download(url: str, dest=None):
    import urllib.request

    dest = dest or RAW / url.rsplit("/", 1)[-1]
    if dest.exists():
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=180) as resp, dest.open("wb") as fh:
        fh.write(resp.read())
    return dest


def period_of(value, frequency: str) -> str:
    """Normalise a Date cell to 'YYYY-MM' / 'YYYY-Qn' / 'YYYY'.

    Cell types vary by workbook: monthly sheets carry real dates, quarterly
    sheets carry '2005Q1' strings, annual sheets carry a bare year.
    """
    if frequency == "annual":
        return str(int(num(value)))
    if frequency == "quarterly":
        if isinstance(value, str) and (m := re.fullmatch(r"\s*(\d{4})\s*Q([1-4])\s*", value)):
            return f"{m.group(1)}-Q{m.group(2)}"
        if hasattr(value, "year"):
            return f"{value.year:04d}-Q{(value.month - 1) // 3 + 1}"
    elif hasattr(value, "year"):
        return f"{value.year:04d}-{value.month:02d}"
    raise DecodeError(f"unparseable date cell {value!r} for {frequency}")


def parse_workbook(path, sheet_names, frequency: str, adjusted: bool):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet, label in sheet_names.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}
        if "Date" not in idx:
            raise DecodeError(f"{path.name}:{sheet} has no Date column")

        for raw in rows:
            if not raw or raw[idx["Date"]] is None:
                continue
            period = period_of(raw[idx["Date"]], frequency)
            for crea_type, ptype in TYPES.items():
                hpi = idx.get(f"{crea_type}_HPI")
                bench = idx.get(f"{crea_type}_Benchmark")
                hpi_v = num(raw[hpi]) if hpi is not None and hpi < len(raw) else None
                bench_v = num(raw[bench]) if bench is not None and bench < len(raw) else None
                if hpi_v is None and bench_v is None:
                    continue
                yield {
                    "source": "crea_hpi",
                    "source_file": path.name,
                    "area_raw": label,
                    "area_slug": slugify(label),
                    "crea_sheet": sheet,
                    "property_type": ptype,
                    "frequency": frequency,
                    "seasonally_adjusted": adjusted,
                    "period": period,
                    "hpi": hpi_v,
                    "benchmark_price": bench_v,
                }


def main() -> int:
    url = find_latest_zip_url()
    print(f"CREA archive: {url}")
    zip_path = download(url)

    records = []
    with zipfile.ZipFile(zip_path) as zf:
        names = set(zf.namelist())
        extract_dir = RAW / "crea_hpi"
        extract_dir.mkdir(parents=True, exist_ok=True)
        for member, (frequency, adjusted) in WORKBOOKS.items():
            if member not in names:
                print(f"  ! missing {member} in archive", file=sys.stderr)
                continue
            target = extract_dir / member
            if not target.exists():
                target.write_bytes(zf.read(member))
            records.extend(parse_workbook(target, SHEETS, frequency, adjusted))

    if not records:
        raise DecodeError("parsed zero CREA records")

    monthly = [r for r in records if r["frequency"] == "monthly" and not r["seasonally_adjusted"]]
    periods = sorted({r["period"] for r in monthly})
    path = OUT / "crea_hpi.jsonl"
    n = write_jsonl(path, records)

    print(f"wrote {n:,} records -> {path}")
    print(f"  monthly NSA range : {periods[0]} .. {periods[-1]}  ({len(periods)} months)")
    print(f"  areas             : {sorted({r['area_raw'] for r in records})}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
